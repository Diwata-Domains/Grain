from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from grain.domain import OfficeArtifactRef, OfficeWriteRequest
from grain.services.spreadsheet_write_service import (
    SpreadsheetCellUpdate,
    SpreadsheetWriteService,
)


def _create_xlsx(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue"
    ws["A1"] = "month"
    ws["B1"] = "amount"
    ws["A2"] = "Jan"
    ws["B2"] = 12
    ws["A3"] = "Feb"
    ws["B3"] = "=SUM(B2,18)"
    summary = wb.create_sheet("Summary")
    summary["A1"] = "owner"
    summary["B1"] = "team-a"
    wb.save(path)


def test_spreadsheet_propose_writes_reviewable_output_without_mutating_source(
    tmp_path: Path,
) -> None:
    source = tmp_path / "data" / "report.xlsx"
    output = tmp_path / "tasks" / "P23-T03-TASK-0154" / "report.proposed.xlsx"
    _create_xlsx(source)

    service = SpreadsheetWriteService()
    request = OfficeWriteRequest(
        packet_id="TASK-0154",
        artifact=OfficeArtifactRef(
            "spreadsheet",
            "data/report.xlsx",
            output_path="tasks/P23-T03-TASK-0154/report.proposed.xlsx",
        ),
        requested_mode="propose",
    )

    result = service.write_workbook(
        root=tmp_path,
        request=request,
        updates=[
            SpreadsheetCellUpdate("Revenue", "B2", 14),
            SpreadsheetCellUpdate("Summary", "B1", "team-b"),
        ],
    )

    source_wb = load_workbook(source, data_only=False)
    output_wb = load_workbook(output, data_only=False)

    assert source_wb["Revenue"]["B2"].value == 12
    assert output_wb["Revenue"]["B2"].value == 14
    assert output_wb["Summary"]["B1"].value == "team-b"
    assert result.output_path == "tasks/P23-T03-TASK-0154/report.proposed.xlsx"
    assert result.touched_sheets == ["Revenue", "Summary"]
    assert result.touched_ranges == ["Revenue!B2", "Summary!B1"]
    assert "mode: propose" in result.change_summary


def test_spreadsheet_export_mode_persists_export_file_and_formula_summary(
    tmp_path: Path,
) -> None:
    source = tmp_path / "data" / "report.xlsx"
    output = tmp_path / "exports" / "report.review.xlsx"
    _create_xlsx(source)

    service = SpreadsheetWriteService()
    request = OfficeWriteRequest(
        packet_id="TASK-0154",
        artifact=OfficeArtifactRef(
            "spreadsheet",
            "data/report.xlsx",
            output_path="exports/report.review.xlsx",
        ),
        requested_mode="export-as-new-file",
    )

    result = service.write_workbook(
        root=tmp_path,
        request=request,
        updates=[SpreadsheetCellUpdate("Revenue", "B3", "=SUM(B2,20)")],
    )

    output_wb = load_workbook(output, data_only=False)

    assert output.exists()
    assert output_wb["Revenue"]["B3"].value == "=SUM(B2,20)"
    assert result.formula_cells_changed == ["Revenue!B3"]
    assert "mode: export-as-new-file" in result.change_summary
    assert "formula cells changed: Revenue!B3" in result.change_summary


def test_spreadsheet_apply_mode_is_rejected_for_phase_23(tmp_path: Path) -> None:
    source = tmp_path / "data" / "report.xlsx"
    _create_xlsx(source)

    service = SpreadsheetWriteService()
    request = OfficeWriteRequest(
        packet_id="TASK-0154",
        artifact=OfficeArtifactRef(
            "spreadsheet",
            "data/report.xlsx",
            output_path="data/report.updated.xlsx",
        ),
        requested_mode="apply",
        explicit_apply=True,
        validation_state="passed",
    )

    with pytest.raises(ValueError, match="In-place spreadsheet apply mode is not supported"):
        service.write_workbook(
            root=tmp_path,
            request=request,
            updates=[SpreadsheetCellUpdate("Revenue", "B2", 14)],
        )


def test_spreadsheet_requires_explicit_output_path(tmp_path: Path) -> None:
    source = tmp_path / "data" / "report.xlsx"
    _create_xlsx(source)

    service = SpreadsheetWriteService()
    request = OfficeWriteRequest(
        packet_id="TASK-0154",
        artifact=OfficeArtifactRef("spreadsheet", "data/report.xlsx"),
        requested_mode="propose",
    )

    with pytest.raises(ValueError, match="outputs require OfficeArtifactRef.output_path"):
        service.write_workbook(
            root=tmp_path,
            request=request,
            updates=[SpreadsheetCellUpdate("Revenue", "B2", 14)],
        )


def test_spreadsheet_update_requires_sheet_name() -> None:
    with pytest.raises(ValueError, match="sheet_name must not be empty"):
        SpreadsheetCellUpdate("", "A1", "x")
