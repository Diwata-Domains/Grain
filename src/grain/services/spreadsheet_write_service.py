"""Writable spreadsheet workflow support for Phase 23 office artifacts."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from grain.domain.office_writes import OfficeWriteDecision, OfficeWriteRequest
from grain.services.office_write_service import OfficeWriteService


@dataclass(frozen=True)
class SpreadsheetCellUpdate:
    sheet_name: str
    cell_ref: str
    value: object

    def __post_init__(self) -> None:
        if not self.sheet_name.strip():
            raise ValueError("SpreadsheetCellUpdate.sheet_name must not be empty")
        if not self.cell_ref.strip():
            raise ValueError("SpreadsheetCellUpdate.cell_ref must not be empty")


@dataclass(frozen=True)
class SpreadsheetWriteResult:
    decision: OfficeWriteDecision
    output_path: str
    change_summary: list[str] = field(default_factory=list)
    touched_sheets: list[str] = field(default_factory=list)
    touched_ranges: list[str] = field(default_factory=list)
    formula_cells_changed: list[str] = field(default_factory=list)


class SpreadsheetWriteService:
    """Apply bounded spreadsheet updates via explicit proposal/export outputs."""

    def __init__(self, office_write_service: OfficeWriteService | None = None) -> None:
        self._office_write_service = office_write_service or OfficeWriteService()

    def write_workbook(
        self,
        *,
        root: Path,
        request: OfficeWriteRequest,
        updates: list[SpreadsheetCellUpdate],
    ) -> SpreadsheetWriteResult:
        decision = self._office_write_service.resolve_write_mode(request)
        if decision.resolved_mode == "apply":
            raise ValueError("In-place spreadsheet apply mode is not supported in Phase 23")

        source_path = self._resolve(root, request.artifact.source_path)
        output_path = self._resolve_output(root, request.artifact.output_path)
        workbook = self._load_workbook(source_path)

        touched_sheets: list[str] = []
        touched_ranges: list[str] = []
        formula_cells_changed: list[str] = []

        for update in updates:
            worksheet = workbook[update.sheet_name]
            cell = worksheet[update.cell_ref]
            previous_value = cell.value
            cell.value = update.value

            if update.sheet_name not in touched_sheets:
                touched_sheets.append(update.sheet_name)
            touched_ranges.append(f"{update.sheet_name}!{update.cell_ref}")
            if isinstance(previous_value, str) and previous_value.startswith("="):
                formula_cells_changed.append(f"{update.sheet_name}!{update.cell_ref}")

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)

        change_summary = [
            f"mode: {decision.resolved_mode}",
            f"touched sheets: {', '.join(touched_sheets) if touched_sheets else '(none)'}",
            f"touched ranges: {', '.join(touched_ranges) if touched_ranges else '(none)'}",
            (
                "formula cells changed: "
                f"{', '.join(formula_cells_changed) if formula_cells_changed else '(none)'}"
            ),
        ]
        for update in updates:
            change_summary.append(
                f"cell update: {update.sheet_name}!{update.cell_ref} -> {update.value!r}"
            )

        return SpreadsheetWriteResult(
            decision=decision,
            output_path=str(output_path.relative_to(root)),
            change_summary=change_summary,
            touched_sheets=touched_sheets,
            touched_ranges=touched_ranges,
            formula_cells_changed=formula_cells_changed,
        )

    def _resolve(self, root: Path, raw_path: str) -> Path:
        path = Path(raw_path)
        return path if path.is_absolute() else root / path

    def _resolve_output(self, root: Path, raw_path: str) -> Path:
        if not raw_path.strip():
            raise ValueError(
                "spreadsheet write outputs require OfficeArtifactRef.output_path"
            )
        return self._resolve(root, raw_path)

    def _load_workbook(self, path: Path):
        from openpyxl import load_workbook

        return load_workbook(path)
