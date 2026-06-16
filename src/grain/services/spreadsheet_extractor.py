# SPDX-FileCopyrightText: 2024-2026 Shaznay Sison
# SPDX-License-Identifier: AGPL-3.0-only

"""Spreadsheet extraction service for adapter-driven context assembly."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any


class SpreadsheetExtractor:
    """Extract readable text from spreadsheet files."""

    def extract(self, path: Path) -> str:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return self._extract_csv(path)
        if suffix in {".xlsx", ".xls"}:
            return self._extract_excel(path)
        return f"[spreadsheet_extractor: unsupported file type {path.suffix or '(none)'}]"

    def _extract_csv(self, path: Path) -> str:
        try:
            with path.open("r", encoding="utf-8", newline="") as handle:
                rows = list(csv.reader(handle))
        except Exception as exc:  # noqa: BLE001 - must degrade gracefully for context export
            return f"[spreadsheet_extractor: could not read {path.name} - {exc}]"

        if not rows:
            return f"[spreadsheet_extractor: {path.name} is empty]"

        header = rows[0]
        data_rows = rows[1:]
        lines = [
            f"# Spreadsheet: {path.name}",
            "## Sheet: CSV",
            f"- Columns: {', '.join(self._stringify_cells(header)) if header else '(none)'}",
            f"- Rows: {len(data_rows)}",
        ]
        if data_rows:
            lines.append("## Data")
            for idx, row in enumerate(data_rows, start=1):
                lines.append(f"{idx}. {', '.join(self._stringify_cells(row))}")
        return "\n".join(lines)

    def _extract_excel(self, path: Path) -> str:
        try:
            workbook = self._load_workbook(path)
        except Exception as exc:  # noqa: BLE001 - must degrade gracefully for context export
            return f"[spreadsheet_extractor: could not read {path.name} - {exc}]"

        sheet_names = list(getattr(workbook, "sheetnames", []))
        if not sheet_names:
            return f"[spreadsheet_extractor: {path.name} is empty]"

        lines = [f"# Spreadsheet: {path.name}"]
        for sheet_name in sheet_names:
            worksheet = workbook[sheet_name]
            rows = list(worksheet.iter_rows(values_only=True))
            lines.append("")
            lines.append(f"## Sheet: {sheet_name}")
            if not rows:
                lines.append("- Columns: (none)")
                lines.append("- Rows: 0")
                continue

            header = rows[0] or ()
            data_rows = rows[1:]
            lines.append(f"- Columns: {', '.join(self._stringify_cells(header)) if header else '(none)'}")
            lines.append(f"- Rows: {len(data_rows)}")
            if data_rows:
                lines.append("### Data")
                for idx, row in enumerate(data_rows, start=1):
                    row_text = ", ".join(self._stringify_cells(row or ()))
                    lines.append(f"{idx}. {row_text}")

        return "\n".join(lines)

    def _load_workbook(self, path: Path) -> Any:
        from openpyxl import load_workbook

        return load_workbook(path, read_only=True, data_only=True)

    def _stringify_cells(self, row: Any) -> list[str]:
        values: list[str] = []
        for cell in row:
            if cell is None:
                values.append("")
            else:
                values.append(str(cell))
        return values
